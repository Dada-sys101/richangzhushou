from __future__ import annotations

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
FIXTURES.mkdir(parents=True, exist_ok=True)

WECHAT = ["交易时间", "交易类型", "交易对方", "商品", "收/支", "金额(元)", "支付方式", "当前状态", "交易单号", "商户单号", "备注"]
ALIPAY = ["交易号", "商家订单号", "交易创建时间", "付款时间", "最近修改时间", "交易来源地", "类型", "交易对方", "商品名称", "金额（元）", "收/支", "交易状态", "服务费（元）", "成功退款（元）", "备注", "资金状态"]

with (FIXTURES / "wechat-representative.csv").open("w", encoding="utf-8-sig", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(["微信支付账单明细", "", ""])
    writer.writerow(["导出时间：2026-08-08", "", ""])
    writer.writerow(WECHAT)
    writer.writerow(["2026-08-01 12:30:01", "商户消费", "示例餐厅", "午餐", "支出", "¥12.80", "零钱", "支付成功", "4200001", "M001", ""])
    writer.writerow(["2026-08-02 09:00:00", "转账", "示例用户", "退款", "收入", "3.50", "零钱", "已收钱", "4200002", "", None])

with (FIXTURES / "alipay-representative.csv").open("w", encoding="gb18030", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(["支付宝交易记录明细查询", "", ""])
    writer.writerow(ALIPAY)
    writer.writerow(["202608010001", "A001", "2026-08-01 12:35:00", "2026-08-01 12:35:05", "2026-08-01 12:35:05", "支付宝", "即时到账交易", "示例餐厅", "午餐", "12.80", "支出", "交易成功", "0.00", "0.00", "", "已支出"])
    writer.writerow(["202608020001", "A002", "2026-08-02 09:01:00", "", "2026-08-02 09:10:00", "支付宝", "退款", "示例商户", "退款", "3.50", "收入", "退款成功", "0.00", "3.50", None, "已收入"])

source = FIXTURES / "alipay-representative.csv"
(FIXTURES / "alipay-representative-utf8.csv").write_text(source.read_text(encoding="gb18030"), encoding="utf-8-sig")

book = Workbook()
sheet = book.active
sheet.title = "账单"
sheet.merge_cells("A1:P1")
sheet["A1"] = "支付宝交易记录明细查询"
sheet["A1"].font = Font(bold=True)
sheet.append(ALIPAY)
sheet.append(["202608010001", "A001", "2026-08-01 12:35:00", "2026-08-01 12:35:05", "2026-08-01 12:35:05", "支付宝", "即时到账交易", "示例餐厅", "午餐", 12.8, "支出", "交易成功", 0, 0, None, "已支出"])
sheet.append(["202608020001", "A002", "2026-08-02 09:01:00", None, "2026-08-02 09:10:00", "支付宝", "退款", "示例商户", "退款", 3.5, "收入", "退款成功", 0, 3.5, "", "已收入"])
book.save(FIXTURES / "alipay-representative.xlsx")

with (FIXTURES / "large-100k.csv").open("w", encoding="utf-8-sig", newline="") as file:
    writer = csv.writer(file)
    writer.writerow(WECHAT)
    for index in range(100_000):
        writer.writerow(["2026-08-01 12:30:01", "商户消费", f"商户{index % 50}", "测试商品", "支出", f"{(index % 5000) / 100:.2f}", "零钱", "支付成功", f"420{index:012d}", f"M{index:012d}", ""])

book = Workbook(write_only=True)
sheet = book.create_sheet("账单")
sheet.append(ALIPAY)
for index in range(100_000):
    sheet.append([f"2026{index:012d}", f"A{index:012d}", "2026-08-01 12:35:00", "2026-08-01 12:35:05", "2026-08-01 12:35:05", "支付宝", "即时到账交易", f"商户{index % 50}", "测试商品", (index % 5000) / 100, "支出", "交易成功", 0, 0, None, "已支出"])
book.save(FIXTURES / "large-100k.xlsx")

print({path.name: path.stat().st_size for path in FIXTURES.iterdir()})
